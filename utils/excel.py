from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from utils.logger import logger
from utils.schemas import ParsedProject, ProjectComment, ProjectTask, WorkbookIssue


def _norm(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _to_bool(value: Any) -> bool | None:
    if pd.isna(value):
        return None
    text = str(value).strip().lower()
    if text in {"yes", "y", "true", "1", "blocked", "risk", "at risk"}:
        return True
    if text in {"no", "n", "false", "0", ""}:
        return False
    return None


def _to_date(value: Any) -> datetime | None:
    if value is None or pd.isna(value):
        return None
    try:
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.to_pydatetime()
    except Exception:
        return None


def _to_percent(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.strip().replace("%", "")
        if not value or value.upper() == "#UNPARSEABLE":
            return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number <= 1:
        return round(number * 100, 2)
    return round(number, 2)


def _to_int(value: Any) -> int | None:
    if value is None or pd.isna(value):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _parse_variance_days(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip().lower()
    if not text or text == "#unparseable":
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


def _find_header_row(raw: pd.DataFrame) -> int:
    best_index = 0
    best_score = -1
    required_terms = {"task", "status", "complete", "health", "date", "manager", "milestone"}
    for index, row in raw.head(15).iterrows():
        values = {_norm(v) for v in row.tolist() if str(v).strip()}
        score = sum(any(term in value for value in values) for term in required_terms)
        if score > best_score:
            best_score = score
            best_index = index
    return int(best_index)


def _load_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None, engine="openpyxl")
    header_row = _find_header_row(raw)
    headers = [str(v).strip() if not pd.isna(v) else f"column_{i}" for i, v in enumerate(raw.iloc[header_row])]
    df = raw.iloc[header_row + 1 :].copy()
    df.columns = headers
    df = df.dropna(how="all")
    return df


def discover_workbook(path: str | Path) -> dict[str, Any]:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        sample = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            sample.append([cell for cell in row[:12]])
        sheets.append({"name": ws.title, "rows": ws.max_row, "columns": ws.max_column, "sample": sample})
    return {"path": str(workbook_path), "sheets": sheets}


def _column_map(columns: list[str]) -> dict[str, str]:
    aliases = {
        "task_name": ["task name", "task", "activity", "name"],
        "status": ["status", "task status"],
        "start_date": ["start date", "baseline start date", "target start", "planned start"],
        "end_date": ["end date", "baseline end date", "finish date", "target end", "planned end"],
        "percent_complete": ["complete", "% complete", "completion"],
        "schedule_health": ["schedule health", "health"],
        "phase_milestone": ["phase milestone", "milestone", "phase"],
        "project_manager": ["project manager", "manager"],
        "at_risk": ["at risk", "risk"],
        "on_hold": ["on hold", "hold"],
        "critical": ["critical"],
        "blocked": ["blocked", "blocker"],
        "variance_days": ["variance", "delay", "slippage"],
        "dependencies": ["dependency", "predecessor", "successor"],
        "rag": ["rag"],
        "assigned_to": ["assigned to", "owner"],
        "baseline_start": ["baseline start"],
        "baseline_finish": ["baseline finish"],
    }
    normalized = {_norm(col): col for col in columns}
    mapping: dict[str, str] = {}
    for canonical, choices in aliases.items():
        for choice in choices:
            for norm_col, original in normalized.items():
                if norm_col == choice:
                    mapping[canonical] = original
                    break
            if canonical in mapping:
                break
            for norm_col, original in normalized.items():
                if choice in norm_col:
                    mapping[canonical] = original
                    break
            if canonical in mapping:
                break
    return mapping


def _choose_plan_sheet(path: Path, sheet_names: list[str]) -> tuple[str | None, list[WorkbookIssue]]:
    issues: list[WorkbookIssue] = []
    scored: list[tuple[int, str]] = []
    for sheet in sheet_names:
        try:
            df = _load_sheet(path, sheet)
        except Exception as exc:
            issues.append(WorkbookIssue(level="warning", message=f"Could not inspect sheet {sheet}: {exc}"))
            continue
        mapping = _column_map([str(c) for c in df.columns])
        score = len(set(mapping) & {"task_name", "status", "percent_complete", "schedule_health"})
        if _norm(sheet) in {"comments", "summary"}:
            score -= 4
        scored.append((score, sheet))
    scored.sort(reverse=True)
    if not scored or scored[0][0] <= 0:
        return None, issues + [WorkbookIssue(level="error", message="No project plan sheet could be inferred.")]
    return scored[0][1], issues


def _parse_summary(path: Path, sheet_name: str | None) -> tuple[dict[str, Any], list[WorkbookIssue]]:
    if not sheet_name:
        return {}, []
    try:
        raw = pd.read_excel(path, sheet_name=sheet_name, header=None, engine="openpyxl")
    except Exception as exc:
        return {}, [WorkbookIssue(level="warning", message=f"Could not parse summary sheet: {exc}")]
    summary: dict[str, Any] = {}
    for _, row in raw.iterrows():
        key = row.iloc[0] if len(row) else None
        value = row.iloc[1] if len(row) > 1 else None
        if key is not None and not pd.isna(key):
            summary[str(key).strip()] = None if pd.isna(value) else value
    return summary, []


def _parse_comments(path: Path, sheet_name: str | None) -> tuple[list[ProjectComment], list[WorkbookIssue]]:
    if not sheet_name:
        return [], []
    try:
        raw = pd.read_excel(path, sheet_name=sheet_name, header=None, engine="openpyxl")
    except Exception as exc:
        return [], [WorkbookIssue(level="warning", message=f"Could not parse comments sheet: {exc}")]
    comments: list[ProjectComment] = []
    for _, row in raw.dropna(how="all").iterrows():
        values = [None if pd.isna(v) else v for v in row.tolist()]
        text = str(values[1]).strip() if len(values) > 1 and values[1] is not None else ""
        if not text:
            continue
        comments.append(
            ProjectComment(
                row_reference=str(values[0] or ""),
                text=text,
                author=str(values[2] or "") if len(values) > 2 else "",
                created_at=_to_date(values[3]) if len(values) > 3 else None,
            )
        )
    return comments, []


def parse_project_workbook(path: str | Path) -> ParsedProject:
    workbook_path = Path(path)
    issues: list[WorkbookIssue] = []
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        return ParsedProject(
            source_file=str(workbook_path),
            parsed_at=datetime.now(),
            issues=[WorkbookIssue(level="error", message=f"Could not open workbook: {exc}")],
        )

    sheet_names = wb.sheetnames
    plan_sheet, plan_issues = _choose_plan_sheet(workbook_path, sheet_names)
    issues.extend(plan_issues)
    summary_sheet = next((s for s in sheet_names if "summary" in _norm(s)), None)
    comments_sheet = next((s for s in sheet_names if "comment" in _norm(s)), None)
    summary, summary_issues = _parse_summary(workbook_path, summary_sheet)
    comments, comment_issues = _parse_comments(workbook_path, comments_sheet)
    issues.extend(summary_issues + comment_issues)

    tasks: list[ProjectTask] = []
    if plan_sheet:
        try:
            df = _load_sheet(workbook_path, plan_sheet)
            mapping = _column_map([str(c) for c in df.columns])
            missing = [name for name in ["task_name", "status", "percent_complete"] if name not in mapping]
            for name in missing:
                issues.append(WorkbookIssue(level="warning", message=f"Missing expected column: {name}"))
            for row_index, row in df.iterrows():
                raw = {str(k): (None if pd.isna(v) else v) for k, v in row.to_dict().items()}
                task_name = str(raw.get(mapping.get("task_name", ""), "") or "").strip()
                if not task_name and not raw:
                    continue
                variance = _parse_variance_days(raw.get(mapping.get("variance_days", "")))
                tasks.append(
                    ProjectTask(
                        row_number=int(row_index) + 2,
                        task_name=task_name,
                        status=str(raw.get(mapping.get("status", ""), "") or "").strip(),
                        start_date=_to_date(raw.get(mapping.get("start_date", ""))),
                        end_date=_to_date(raw.get(mapping.get("end_date", ""))),
                        percent_complete=_to_percent(raw.get(mapping.get("percent_complete", ""))),
                        schedule_health=str(raw.get(mapping.get("schedule_health", ""), "") or "").strip(),
                        phase_milestone=str(raw.get(mapping.get("phase_milestone", ""), "") or "").strip(),
                        project_manager=str(raw.get(mapping.get("project_manager", ""), "") or "").strip(),
                        at_risk=_to_bool(raw.get(mapping.get("at_risk", ""))),
                        on_hold=_to_bool(raw.get(mapping.get("on_hold", ""))),
                        critical=_to_bool(raw.get(mapping.get("critical", ""))),
                        blocked=_to_bool(raw.get(mapping.get("blocked", ""))),
                        variance_days=variance,
                        dependencies=str(raw.get(mapping.get("dependencies", ""), "") or "").strip(),
                        rag=str(raw.get(mapping.get("rag", ""), "") or "").strip(),
                        assigned_to=str(raw.get(mapping.get("assigned_to", ""), "") or "").strip(),
                        baseline_start=_to_date(raw.get(mapping.get("baseline_start", ""))),
                        baseline_finish=_to_date(raw.get(mapping.get("baseline_finish", ""))),
                        raw=raw,
                    )
                )
        except Exception as exc:
            logger.exception("Plan parsing failed")
            issues.append(WorkbookIssue(level="error", message=f"Could not parse plan sheet: {exc}"))

    first_task = tasks[0] if tasks else None
    summary_counts: dict[str, int] = {}
    for key in ["Not Started", "In Progress", "Completed", "On Hold"]:
        parsed_count = _to_int(summary.get(key))
        if parsed_count is not None:
            summary_counts[key] = parsed_count
    project_name = str(summary.get("Project Name") or (first_task.task_name if first_task else "Unknown Project")).strip()
    return ParsedProject(
        source_file=str(workbook_path),
        parsed_at=datetime.now(),
        plan_sheet=plan_sheet,
        summary_sheet=summary_sheet,
        comments_sheet=comments_sheet,
        project_name=project_name or "Unknown Project",
        project_manager=str(summary.get("Project Manager") or (first_task.project_manager if first_task else "")),
        project_stage=str(summary.get("Project Stage") or ""),
        project_status=str(summary.get("Project Status") or ""),
        schedule_health=str(summary.get("Schedule Health") or ""),
        summary_completion_percent=_to_percent(summary.get("% Complete")),
        summary_task_counts=summary_counts,
        project_start_date=_to_date(summary.get("Project Start Date")),
        project_end_date=_to_date(summary.get("Project End Date")),
        summary=summary,
        tasks=tasks,
        comments=comments,
        issues=issues,
    )
